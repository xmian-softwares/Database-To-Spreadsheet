import mysql.connector
import openpyxl
import webbrowser
from openpyxl.styles import PatternFill, Color, Fill, Font
import os
from datetime import date
from datetime import datetime

def stud1(last):
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()
  sheet = wb["Sheet"]
  sheet.column_dimensions["B"].width = "50"
  sheet.column_dimensions["A"].width = "17"
  sheet.column_dimensions["C"].width = "17"
  sheet.column_dimensions["D"].width = "17"
  sheet.column_dimensions["E"].width = "17"

  mycursor = mydb.cursor()
  query = "SELECT * FROM registery WHERE `Class`='"+str(last)+"' AND `Active`='0'"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "ID"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["B1"] = "Full Name"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Class"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["D1"] = "Sex"
  sheet['D1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['D1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["E1"] = "Shift"
  sheet['E1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['E1'].font = Font(color = "FFFFFF",bold= True,size=14)

  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
    D = "D"+str(start)
    E = "E"+str(start)

    sheet[A] = "MW"+str(x[0])
    sheet[B] = str(x[1])
    sheet[C] = str(x[9])
    sheet[D] = str(x[10])
    sheet[E] = str(x[15])

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//"+str(last)+" "+str(now)+str(time)+".xlsx")

def stud2():
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()
  sheet = wb["Sheet"]
  sheet.column_dimensions["B"].width = "50"

  mycursor = mydb.cursor()
  query = "SELECT * FROM registery WHERE `Active`='0'"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "ID"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["B1"] = "Full Name"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Class"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["D1"] = "Sex"
  sheet['D1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['D1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["E1"] = "Shift"
  sheet['E1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['E1'].font = Font(color = "FFFFFF",bold= True,size=14)

  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
    D = "D"+str(start)
    E = "E"+str(start)

    sheet[A] = "MW"+str(x[0])
    sheet[B] = str(x[1])
    sheet[C] = str(x[9])+" "+str(x[17])
    sheet[D] = str(x[10])
    sheet[E] = str(x[15])

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//All Students List "+str(now)+str(time)+".xlsx")
def stud3(last):
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()
  sheet = wb["Sheet"]
  sheet.column_dimensions["B"].width = "50"
  sheet.column_dimensions["A"].width = "17"
  sheet.column_dimensions["C"].width = "17"
  sheet.column_dimensions["D"].width = "17"

  mycursor = mydb.cursor()
  query = "SELECT * FROM `average` WHERE `Class`='"+str(last)+"'"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "ID"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["B1"] = "Full Name"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Class"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["D1"] = "Average"
  sheet['D1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['D1'].font = Font(color = "FFFFFF",bold= True,size=14)

  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
    D = "D"+str(start)

    sheet[A] = "MW"+str(x[0])
    sheet[B] = str(x[1])
    sheet[C] = str(x[2])
    sheet[D] = str(x[3])+"%"

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//"+str(last)+" Result "+str(now)+str(time)+".xlsx")

def stud4():
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()
  sheet = wb["Sheet"]
  sheet.column_dimensions["B"].width = "50"
  sheet.column_dimensions["A"].width = "17"
  sheet.column_dimensions["C"].width = "17"
  sheet.column_dimensions["D"].width = "17"

  mycursor = mydb.cursor()
  query = "SELECT * FROM `average`"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "ID"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["B1"] = "Full Name"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Class"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["D1"] = "Average"
  sheet['D1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['D1'].font = Font(color = "FFFFFF",bold= True,size=14)

  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
    D = "D"+str(start)

    sheet[A] = "MW"+str(x[0])
    sheet[B] = str(x[1])
    sheet[C] = str(x[2])
    sheet[D] = str(x[3])+"%"

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//All Students Result List "+str(now)+str(time)+".xlsx")
def stud5():
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()
  sheet = wb["Sheet"]
  sheet.column_dimensions["B"].width = "20"
  sheet.column_dimensions["A"].width = "30"
  sheet.column_dimensions["C"].width = "20"
  sheet.column_dimensions["D"].width = "20"
  sheet.column_dimensions["E"].width = "20"
  sheet.column_dimensions["F"].width = "20"
  sheet.column_dimensions["G"].width = "20"

  mycursor = mydb.cursor()
  query = "SELECT * FROM `day_rep` ORDER BY `Id` DESC"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "Day"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["B1"] = "Registered"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Paid"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["D1"] = "Stopped"
  sheet['D1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['D1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["E1"] = "Rejoined"
  sheet['E1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['E1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["F1"] = "Total Income"
  sheet['F1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['F1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["G1"] = "Total Expense"
  sheet['G1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['G1'].font = Font(color = "FFFFFF",bold= True,size=14)

  payn = 0
  payrn = 0
  pays = 0
  payrs = 0
  paysp = 0

  query2 = "SELECT * FROM `money`"
  mycursor.execute(query2)

  myresult2 = mycursor.fetchall()
  for y in myresult2:
    payn = int(y[1])
    payrn = int(y[2])
    pays = int(y[3])
    payrs = int(y[4])
    
  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
    D = "D"+str(start)
    E = "E"+str(start)
    F = "F"+str(start)
    G = "G"+str(start)

    tot_exp = 0
    exday = str(x[1])
    exmon = str(x[2])
    exyea = str(x[3])
    query3 = "SELECT * FROM `expense` WHERE `Day`='"+str(exday)+"' AND `Month`='"+str(exmon)+"' AND `Year`='"+str(exyea)+"'"
    mycursor.execute(query3)

    myresult3 = mycursor.fetchall()
    for z in myresult3:
      tot_exp = tot_exp + int(z[5])

    sheet[A] = str(x[1])+"-"+str(x[2])+"-"+str(x[3])
    sheet[B] = str(int(x[5])+int(x[8])+int(x[10]))+" Students"
    sheet[C] = str(int(x[4])+int(x[9])+int(x[11]))+" Students"
    sheet[D] = str(x[7])+" Students"
    sheet[E] = str(x[6])+" Students"
    
    sheet[F] = str((int(x[4])*payn)+(int(x[9])*pays)+(int(x[11])*paysp)+(int(x[5])*payrn)+(int(x[8])*payrs)+(int(x[10])*paysp)+(int(x[6])*payrn))+" Birr"
    sheet[G] = str(tot_exp)+" Birr"

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//Daily Reports "+str(now)+str(time)+".xlsx")

def stud6():
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()

  sheet = wb["Sheet"]
  sheet.column_dimensions["B"].width = "20"
  sheet.column_dimensions["A"].width = "30"
  sheet.column_dimensions["C"].width = "20"
  sheet.column_dimensions["D"].width = "20"
  sheet.column_dimensions["E"].width = "20"
  sheet.column_dimensions["F"].width = "20"

  mycursor = mydb.cursor()
  query = "SELECT * FROM `mon_rep` ORDER BY `Id` DESC"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "Month"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["B1"] = "Registered"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Stopped"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
  
  sheet["D1"] = "Rejoined"
  sheet['D1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['D1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["E1"] = "Total Income"
  sheet['E1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['E1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["F1"] = "Total Expense"
  sheet['F1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['F1'].font = Font(color = "FFFFFF",bold= True,size=14)

  payn = 0
  payrn = 0
  pays = 0
  payrs = 0
  paysp = 0

  query2 = "SELECT * FROM `money`"
  mycursor.execute(query2)

  myresult2 = mycursor.fetchall()
  for y in myresult2:
    payn = int(y[1])
    payrn = int(y[2])
    pays = int(y[3])
    payrs = int(y[4])
    
  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
    D = "D"+str(start)
    E = "E"+str(start)
    F = "F"+str(start)

    exmon = str(x[1])
    exyea = str(x[2])

    query6 = "SELECT * FROM `day_rep` WHERE `Month`='"+str(exmon)+"' AND `Year`='"+str(exyea)+"' ORDER BY `Id` DESC"
    mycursor.execute(query6)

    myresult6 = mycursor.fetchall()
    tot = 0
    for j in myresult6:
      tot = tot +(int(j[4])*payn)+(int(j[9])*pays)

    query7 = "SELECT * FROM `expense` WHERE `Month`='"+str(exmon)+"' AND `Year`='"+str(exyea)+"' ORDER BY `Id` DESC"
    mycursor.execute(query7)

    myresult7 = mycursor.fetchall()
    tot2 = 0
    for k in myresult7:
      tot2 = tot2 +int(k[5])
      
    
    sheet[A] = "Month "+str(x[0])+" ("+str(x[1])+"-"+str(x[2])+")"
    sheet[B] = str(int(x[4])+int(x[5])+int(x[6]))+" Students"
    sheet[C] = str(x[8])+" Students"
    sheet[D] = str(x[7])+" Students"
    sheet[E] = str((int(x[4])*payrn)+(int(x[5])*payrs)+(int(x[6])*paysp)+(int(x[7])*payrn)+tot)+" Birr"
    sheet[F] = str(tot2)+" Birr"

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//Monthly Reports "+str(now)+str(time)+".xlsx")
def stud7():
  webbrowser.open('http://localhost/multiway')

def stud8():
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()

  sheet = wb["Sheet"]
  sheet.column_dimensions["A"].width = "30"
  sheet.column_dimensions["B"].width = "50"
  sheet.column_dimensions["C"].width = "20"

  mycursor = mydb.cursor()
  query = "SELECT * FROM `expense` ORDER BY `Id` DESC"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "Date"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["B1"] = "Expense"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Amount"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
    
  for x in myresult:
    month = date.today().strftime("%m")
    year = date.today().strftime("%Y")

    nyear = str(int(year)-1)
    nmonth = int(month)
    sm = int(x[2])

    if month == "01":
      if x[2]== month and x[3]== year:
        A = "A"+str(start)
        B = "B"+str(start)
        C = "C"+str(start)
          
        
        sheet[A] = str(x[1])+"-"+str(x[2])+"-"+str(x[3])
        sheet[B] = str(x[4])
        sheet[C] = str(x[5])+" Birr"

        start = start+1
      if x[2] == "12" and x[3] == nyear:
        A = "A"+str(start)
        B = "B"+str(start)
        C = "C"+str(start)
          
        
        sheet[A] = str(x[1])+"-"+str(x[2])+"-"+str(x[3])
        sheet[B] = str(x[4])
        sheet[C] = str(x[5])+" Birr"

        start = start+1

    else:
      if x[2]== month and x[3]== year:
        A = "A"+str(start)
        B = "B"+str(start)
        C = "C"+str(start)
          
        
        sheet[A] = str(x[1])+"-"+str(x[2])+"-"+str(x[3])
        sheet[B] = str(x[4])
        sheet[C] = str(x[5])+" Birr"

        start = start+1
      if sm == nmonth and x[3] == year:
        A = "A"+str(start)
        B = "B"+str(start)
        C = "C"+str(start)
          
        
        sheet[A] = str(x[1])+"-"+str(x[2])+"-"+str(x[3])
        sheet[B] = str(x[4])
        sheet[C] = str(x[5])+" Birr"

        start = start+1

  now2 = datetime.now()
  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//Recent Expenses "+str(now)+str(time)+".xlsx")

def stud9():
  mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="",
    database="register"
  )
 
  wb = openpyxl.Workbook()

  sheet = wb["Sheet"]
  sheet.column_dimensions["A"].width = "30"
  sheet.column_dimensions["B"].width = "50"
  sheet.column_dimensions["C"].width = "20"

  mycursor = mydb.cursor()
  query = "SELECT * FROM `expense` ORDER BY `Id` DESC"
  mycursor.execute(query)

  myresult = mycursor.fetchall()

  start = 2
  sheet["A1"] = "Date"
  sheet['A1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['A1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["B1"] = "Expense"
  sheet['B1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['B1'].font = Font(color = "FFFFFF",bold= True,size=14)

  sheet["C1"] = "Amount"
  sheet['C1'].fill = PatternFill(bgColor="000000", fill_type = "solid")
  sheet['C1'].font = Font(color = "FFFFFF",bold= True,size=14)
    
  for x in myresult:
    A = "A"+str(start)
    B = "B"+str(start)
    C = "C"+str(start)
      
    
    sheet[A] = str(x[1])+"-"+str(x[2])+"-"+str(x[3])
    sheet[B] = str(x[4])
    sheet[C] = str(x[5])+" Birr"

    start = start+1

  now2 = datetime.now()

  time = now2.strftime("%H%M%S")
  now = date.today()

  wb.save("Saved//All Expenses "+str(now)+str(time)+".xlsx")


  
